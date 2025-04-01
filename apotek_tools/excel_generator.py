"""Module for generating Excel files from drug data."""

import os
from typing import Dict, List, Any, Optional

import pandas as pd
from rich.console import Console
from openpyxl.styles import Font, Alignment, Border, Side

from apotek_tools.fetcher import get_current_date
from apotek_tools.config import get_contact_info

console = Console()


def generate_excel(drugs: List[Dict[str, Any]], output_file: Optional[str] = None, config_file: Optional[str] = None) -> str:
    """Generate an Excel file from the drug data.
    
    Args:
        drugs (List[Dict[str, Any]]): Processed drug list
        output_file (Optional[str], optional): Output file path. Defaults to None.
        config_file (Optional[str], optional): Path to config file. Defaults to None.
    
    Returns:
        str: Path to the generated Excel file
    """
    if not drugs:
        console.print("No drug data to export!", style="bold red")
        return ""
    
    # Get contact information from config
    contact_info = get_contact_info(config_file)
    
    # Create a DataFrame from the drug data
    df = pd.DataFrame([{
        "Nama Obat": drug["name"],
        "Harga Diskon": drug["discount_price"],
        "Sisa Stok": drug["stock"]
    } for drug in drugs])
    
    # Generate a default filename if none is provided
    if output_file is None:
        current_date = get_current_date().replace(" ", "_")
        output_file = f"Daftar_Harga_Apotek_Aulia_Farma_{current_date}.xlsx"
    
    # Create a writer object
    with pd.ExcelWriter(output_file, engine='openpyxl') as writer:
        # Write the DataFrame to the Excel file, but shift it down to make room for the title and contact info
        # Start the dataframe at row 5 (after title and contact info)
        df.to_excel(writer, sheet_name='Daftar Harga', index=False, startrow=5)
        
        # Get the workbook and the worksheet
        workbook = writer.book
        worksheet = writer.sheets['Daftar Harga']
        
        # Define border styles
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        medium_border = Border(
            left=Side(style='medium'),
            right=Side(style='medium'),
            top=Side(style='medium'),
            bottom=Side(style='medium')
        )
        
        # Add the title
        current_date = get_current_date()
        worksheet.cell(row=1, column=1, value=f"Daftar Harga Apotek Aulia Farma per {current_date}")
        worksheet.merge_cells(start_row=1, start_column=1, end_row=1, end_column=3)
        
        # Style the title
        title_cell = worksheet.cell(row=1, column=1)
        title_cell.font = Font(bold=True, size=14)
        title_cell.alignment = Alignment(horizontal='center')
        
        # Add contact information below the title but above the data
        contact_row = 3  # Row 3 (after title, leaving row 2 blank)
        
        # Add WhatsApp contact
        worksheet.cell(row=contact_row, column=1, value="Kontak WhatsApp:")
        worksheet.cell(row=contact_row, column=2, value=contact_info["whatsapp"])
        
        # Add Email contact in the next row
        worksheet.cell(row=contact_row + 1, column=1, value="Email:")
        worksheet.cell(row=contact_row + 1, column=2, value=contact_info["email"])
        
        # Style the contact info
        for row in range(contact_row, contact_row + 2):
            for col in range(1, 3):  # Columns A, B
                cell = worksheet.cell(row=row, column=col)
                if col == 1:  # Label column
                    cell.font = Font(bold=True)
        
        # Set column widths
        worksheet.column_dimensions['A'].width = 40  # Nama Obat
        worksheet.column_dimensions['B'].width = 30  # Harga Diskon
        worksheet.column_dimensions['C'].width = 20  # Sisa Stok
        
        # Apply text wrapping and borders to header row
        header_row = 6
        for col in range(1, 4):  # Columns A, B, C 
            header_cell = worksheet.cell(row=header_row, column=col)
            header_cell.alignment = Alignment(wrap_text=True, horizontal='center')
            header_cell.font = Font(bold=True)
            header_cell.border = medium_border  # Use medium border for headers
        
        # Apply text wrapping, borders, and adjust row heights for all data cells
        data_start_row = 7  # First row of data after header
        data_end_row = data_start_row + len(drugs) - 1
        
        for row_idx in range(data_start_row, data_end_row + 1):
            for col_idx in range(1, 4):  # Columns A, B, C
                cell = worksheet.cell(row=row_idx, column=col_idx)
                
                # Apply text wrapping to every cell
                cell.alignment = Alignment(wrap_text=True)
                
                # Apply borders to every cell
                cell.border = thin_border
                
                # Calculate row height for cells with newlines
                if cell.value and isinstance(cell.value, str) and '\n' in cell.value:
                    num_lines = cell.value.count('\n') + 1
                    worksheet.row_dimensions[row_idx].height = max(
                        worksheet.row_dimensions[row_idx].height or 15,
                        15 * num_lines  # 15 points per line
                    )
    
    console.print(f"Excel file generated: {output_file}", style="bold green")
    return output_file
